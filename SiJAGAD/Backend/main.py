import os
import time
import bleach
from typing import Optional, List, Dict, Any, Union
from datetime import datetime
from io import BytesIO
from collections import Counter

from fastapi import FastAPI, HTTPException, BackgroundTasks, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from starlette.middleware.base import BaseHTTPMiddleware
from pydantic import BaseModel
from dotenv import load_dotenv
from supabase import create_client, Client
from openpyxl import load_workbook

# --- 0. KONFIGURASI AWAL ---
# Load environment variables
load_dotenv()

# Inisialisasi App
app = FastAPI(
    title="SiJAGAD API",
    description="Sistem Monitoring & Arsip Digital",
    version="1.0.0"
)

# --- 1. SETUP DATABASE (SUPABASE) ---
SUPABASE_URL: str = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY: str = os.getenv("SUPABASE_SERVICE_KEY", "")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("⚠️  WARNING: SUPABASE_URL atau SUPABASE_SERVICE_KEY belum diset di .env")

# Buat Client Supabase
try:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
except Exception as e:
    print(f"❌ Gagal koneksi ke Supabase: {e}")


# --- 2. MODELS (PYDANTIC) ---
class LetterSchema(BaseModel):
    vendor: str
    pekerjaan: str
    nomor_kontrak: str
    tanggal_awal_kontrak: str
    nominal_jaminan: int
    jenis_garansi: str
    nomor_garansi: str
    bank_penerbit: str
    tanggal_awal_garansi: str
    tanggal_akhir_garansi: str
    status: str
    kategori: str
    file_url: Optional[str] = None
    user_email: Optional[str] = "System"
    lokasi: Optional[str] = None


# --- 3. HELPER FUNCTIONS ---
def sanitize_text(text: str) -> str:
    """Membersihkan input text dari tag HTML berbahaya (XSS Protection)."""
    if not text: return ""
    return bleach.clean(text, tags=[], strip=True)

def log_activity_bg(user_email: str, action: str, target: str):
    """Fungsi Background Task untuk mencatat log aktivitas."""
    try:
        supabase.table("activity_logs").insert({
            "user_email": user_email,
            "action": action,
            "target": target,
            "created_at": datetime.now().isoformat()
        }).execute()
        print(f"📝 [Log Saved] {action}: {target}")
    except Exception as e:
        print(f"❌ [Log Error] Gagal mencatat log: {e}")


# --- 4. MIDDLEWARE SECTION ---

# A. Security Headers
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        return response

app.add_middleware(SecurityHeadersMiddleware)

# B. Process Time Monitor
@app.middleware("http")
async def add_process_time_header(request: Request, call_next):
    start_time = time.time()
    response = await call_next(request)
    process_time = time.time() - start_time
    response.headers["X-Process-Time"] = str(process_time)
    return response

# C. CORS (PENTING: Agar Frontend ReValue bisa akses Backend ini)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Mengizinkan semua akses (ubah ke domain spesifik saat production)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# --- 5. ENDPOINTS UTAMA ---

@app.get("/")
def read_root():
    return {
        "status": "online",
        "system": "SiJAGAD API Ready 🚀",
        "timestamp": datetime.now().isoformat()
    }

@app.get("/api/analytics")
def get_analytics_data():
    try:
        # 1. Ambil semua data aktif
        response = supabase.table("letters").select("*").eq("is_deleted", False).execute()
        letters = response.data or []

        total_surat = len(letters)
        total_nominal = 0
        total_expired = 0
        
        status_counts = {} 
        vendor_stats = {}

        for item in letters:
            if not isinstance(item, dict): continue

            raw_status = item.get("status", "Unknown")
            status = str(raw_status).strip()
            status_counts[status] = status_counts.get(status, 0) + 1

            if status.lower() == "expired":
                total_expired += 1

            # Hitung nominal aman
            try:
                nominal_val = item.get("nominal_jaminan", 0)
                # Pastikan jadi int
                nominal = int(float(str(nominal_val))) if nominal_val else 0
            except:
                nominal = 0
            total_nominal += nominal

            raw_vendor = item.get("vendor", "Unknown")
            vendor = str(raw_vendor)
            vendor_stats[vendor] = vendor_stats.get(vendor, 0) + nominal

        # Warna Chart
        color_map = {
            "Aktif": "#10B981",    
            "Baru": "#34D399",     
            "Expired": "#EF4444",  
            "Selesai": "#3B82F6",  
            "Unknown": "#9CA3AF"   
        }

        pie_data = []
        for stat_name, count in status_counts.items():
            pie_data.append({
                "name": stat_name,
                "value": count,
                "color": color_map.get(stat_name, "#9CA3AF") 
            })

        sorted_vendors = sorted(vendor_stats.items(), key=lambda x: x[1], reverse=True)[:5]
        bar_data = [{"name": v[0][:15] + "...", "total": v[1]} for v in sorted_vendors]

        return {
            "summary": {
                "total_surat": total_surat,
                "total_nominal": total_nominal,
                "total_expired": total_expired
            },
            "pie_chart": pie_data,
            "bar_chart": bar_data
        }

    except Exception as e:
        print(f"Analytics Error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/letters/active")
def get_active_letters():
    """Mengambil surat yang MASIH AKTIF."""
    try:
        response = supabase.table("letters").select("*") \
            .eq("is_deleted", False) \
            .neq("status", "Expired") \
            .neq("status", "Selesai") \
            .order("id", desc=True) \
            .execute()
        return response.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database Error: {str(e)}")


@app.get("/letters/archive")
def get_archived_letters():
    """Mengambil surat ARSIP (Expired/Selesai)."""
    try:
        response = supabase.table("letters").select("*") \
            .eq("is_deleted", False) \
            .or_("status.eq.Expired,status.eq.Selesai") \
            .order("id", desc=True) \
            .execute()
        return response.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database Error: {str(e)}")


@app.get("/letters")
def get_all_letters():
    """Mengambil SEMUA surat."""
    try:
        response = supabase.table("letters").select("*") \
            .eq("is_deleted", False) \
            .order("id", desc=True) \
            .execute()
        return response.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/letters")
def create_letter(letter: LetterSchema, background_tasks: BackgroundTasks):
    try:
        data = letter.dict()
        user_email = data.pop("user_email", "Admin")
        data["is_deleted"] = False 
        
        # Sanitasi
        data["vendor"] = sanitize_text(data["vendor"])
        data["pekerjaan"] = sanitize_text(data["pekerjaan"])
        
        response = supabase.table("letters").insert(data).execute()
        
        if response.data:
            background_tasks.add_task(log_activity_bg, user_email, "CREATE", f"Tambah Surat: {data['vendor']}")
        
        return response.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/letters/{letter_id}")
def update_letter(letter_id: int, letter: LetterSchema, background_tasks: BackgroundTasks):
    try:
        data = letter.dict()
        user_email = data.pop("user_email", "Admin")
        
        data["vendor"] = sanitize_text(data["vendor"])
        data["pekerjaan"] = sanitize_text(data["pekerjaan"])
        
        response = supabase.table("letters").update(data).eq("id", letter_id).execute()
        
        if not response.data:
             raise HTTPException(status_code=404, detail="Data tidak ditemukan")

        background_tasks.add_task(log_activity_bg, user_email, "UPDATE", f"Edit Surat: {data['vendor']}")
        return {"status": "success", "data": response.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/letters/{letter_id}")
def delete_letter(letter_id: int, background_tasks: BackgroundTasks, user_email: str = "Admin"): 
    try:
        existing = supabase.table("letters").select("vendor").eq("id", letter_id).execute()
        data_list = existing.data or []
        target_name = "Unknown"
        if len(data_list) > 0 and isinstance(data_list[0], dict):
            target_name = str(data_list[0].get('vendor', 'Unknown'))

        supabase.table("letters").update({"is_deleted": True}).eq("id", letter_id).execute()
        
        background_tasks.add_task(log_activity_bg, user_email, "SOFT_DELETE", f"Hapus Surat: {target_name}")
        return {"status": "success", "message": "Data dipindahkan ke sampah"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# --- 6. SCHEDULER (AUTO EXPIRED) ---
@app.get("/api/cron-update-status")
def cron_auto_update_status(background_tasks: BackgroundTasks):
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        
        response = supabase.table("letters").select("*") \
            .eq("is_deleted", False) \
            .lt("tanggal_akhir_garansi", today) \
            .neq("status", "Expired") \
            .neq("status", "Selesai") \
            .execute()
            
        expired_letters = response.data or []
        if not expired_letters:
            return {"message": "✅ Tidak ada surat yang Expired hari ini."}

        updated_count = 0
        for item in expired_letters:
             if isinstance(item, dict):
                letter_id = item.get('id')
                if letter_id:
                    supabase.table("letters").update({"status": "Expired"}).eq("id", letter_id).execute()
                    updated_count += 1

        log_msg = f"🤖 Auto-Archive: {updated_count} surat diubah menjadi Expired."
        background_tasks.add_task(log_activity_bg, "System CronJob", "AUTO_UPDATE", log_msg)
        return {"status": "success", "message": log_msg, "updated_count": updated_count}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# --- 7. EXPORT EXCEL (PERBAIKAN TYPE CHECKING) ---
@app.get("/export/excel")
def export_excel_template():
    try:
        # Load Template
        template_path = os.path.join(os.path.dirname(__file__), "template_laporan.xlsx")
        
        if not os.path.exists(template_path):
             raise HTTPException(status_code=404, detail="File template_laporan.xlsx tidak ditemukan.")

        wb = load_workbook(template_path)
        ws = wb.active
        
        if ws is None:
            raise HTTPException(status_code=500, detail="Template corrupt: Tidak ada active sheet.")
        
        response = supabase.table("letters").select("*").eq("is_deleted", False).order("id", desc=True).execute()
        data = response.data or []

        row_idx = 5 
        for i, item in enumerate(data, start=1):
            # 1. Cek apakah item benar-benar dictionary
            if not isinstance(item, dict): 
                continue
            
            # 2. TRIK KHUSUS: Definisikan ulang variabel dengan tipe data eksplisit
            # Ini membuat Pylance sadar bahwa 'row_data' PASTI sebuah Dictionary
            row_data: Dict[str, Any] = item

            # Helper function menggunakan 'row_data', BUKAN 'item'
            def get_val(key: str): 
                val = row_data.get(key) # Sekarang .get() aman karena row_data adalah Dict
                return str(val) if val is not None else '-'

            ws[f'A{row_idx}'] = i
            ws[f'B{row_idx}'] = get_val('vendor')
            ws[f'C{row_idx}'] = get_val('nomor_kontrak')
            ws[f'D{row_idx}'] = get_val('pekerjaan')
            
            # Konversi Nominal Aman
            raw_nominal = row_data.get('nominal_jaminan', 0)
            try:
                final_nominal = int(float(str(raw_nominal)))
            except (ValueError, TypeError):
                final_nominal = 0
            
            ws[f'E{row_idx}'] = final_nominal
            ws[f'F{row_idx}'] = get_val('tanggal_akhir_garansi')
            ws[f'G{row_idx}'] = get_val('status')
            
            row_idx += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Laporan_SiJAGAD_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output, 
            headers={'Content-Disposition': f'attachment; filename="{filename}"'},
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"Export Error: {e}")
        raise HTTPException(status_code=500, detail=f"Gagal export: {str(e)}")

@app.get("/logs")
def get_logs():
    try:
        response = supabase.table("activity_logs").select("*").order("created_at", desc=True).limit(50).execute()
        return response.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))